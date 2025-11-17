#!/usr/bin/env python3
"""
Script de generación de calculadoras ORKO CAP-22
Genera 3 archivos .xlsx operativos según CAP22_IMPL_GUIDE.md

Requiere: openpyxl
Instalar: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Colores para headers
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

def create_health_score_calculator(output_path):
    """Genera health_score_calculator.xlsx"""
    wb = Workbook()
    
    # Hoja 1: Inputs
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    ws_inputs['A1'] = "HEALTH SCORE CALCULATOR - INPUTS"
    ws_inputs['A1'].font = Font(bold=True, size=14)
    
    inputs_data = [
        ("Variable", "Valor", "Descripción"),
        ("org_size", 50, "Número de personas"),
        ("context_complexity", 3, "Escala 1-5"),
        ("budget_available", 50000000, "CLP/USD"),
        ("current_state_maturity", 2, "Escala 1-5"),
        ("strategic_clarity", 60, "% (0-100)"),
        ("TF1_capacity_deployed", 40, "% capacidades desplegadas"),
        ("TF2_flow_instrumented", 30, "% flujos instrumentados"),
        ("TF3_info_governed", 25, "% info gobernada"),
    ]
    
    for row_idx, row_data in enumerate(inputs_data, start=2):
        ws_inputs[f'A{row_idx}'] = row_data[0]
        ws_inputs[f'C{row_idx}'] = row_data[1]
        ws_inputs[f'D{row_idx}'] = row_data[2]
        if row_idx == 2:
            ws_inputs[f'A{row_idx}'].fill = HEADER_FILL
            ws_inputs[f'C{row_idx}'].fill = HEADER_FILL
            ws_inputs[f'D{row_idx}'].fill = HEADER_FILL
            ws_inputs[f'A{row_idx}'].font = HEADER_FONT
            ws_inputs[f'C{row_idx}'].font = HEADER_FONT
            ws_inputs[f'D{row_idx}'].font = HEADER_FONT
    
    # Hoja 2: Parametros
    ws_params = wb.create_sheet("Parametros")
    ws_params['A1'] = "PARAMETROS Y PESOS"
    ws_params['A1'].font = Font(bold=True, size=14)
    
    params_data = [
        ("Parámetro", "Valor", "Descripción"),
        ("w_A", 0.33, "Peso Alignment"),
        ("w_P", 0.33, "Peso Performance"),
        ("w_D", 0.34, "Peso Development"),
        ("", "", ""),
        ("TF1_weight", 0.40, "Peso tejido TF1 (Capacity)"),
        ("TF2_weight", 0.35, "Peso tejido TF2 (Flow)"),
        ("TF3_weight", 0.25, "Peso tejido TF3 (Information)"),
        ("", "", ""),
        ("maturity_floor", 0.35, "H_org mínimo viable"),
    ]
    
    for row_idx, row_data in enumerate(params_data, start=2):
        ws_params[f'A{row_idx}'] = row_data[0]
        ws_params[f'C{row_idx}'] = row_data[1]
        ws_params[f'D{row_idx}'] = row_data[2]
        if row_idx == 2:
            ws_params[f'A{row_idx}'].fill = HEADER_FILL
            ws_params[f'C{row_idx}'].fill = HEADER_FILL
            ws_params[f'D{row_idx}'].fill = HEADER_FILL
    
    # Hoja 3: Calculos
    ws_calc = wb.create_sheet("Calculos")
    ws_calc['A1'] = "CALCULOS INTERMEDIOS"
    ws_calc['A1'].font = Font(bold=True, size=14)
    
    ws_calc['A2'] = "Métrica"
    ws_calc['C2'] = "Valor"
    ws_calc['A2'].fill = HEADER_FILL
    ws_calc['C2'].fill = HEADER_FILL
    
    ws_calc['A3'] = "TF1_Score"
    ws_calc['C3'] = '=Inputs!C8/100'
    ws_calc['A4'] = "TF2_Score"
    ws_calc['C4'] = '=Inputs!C9/100'
    ws_calc['A5'] = "TF3_Score"
    ws_calc['C5'] = '=Inputs!C10/100'
    ws_calc['A7'] = "A_Score"
    ws_calc['C7'] = '=(Inputs!C7/100)*0.6 + C3*0.4'
    ws_calc['A8'] = "P_Score"
    ws_calc['C8'] = '=C4*0.7 + (Inputs!C6/5)*0.3'
    ws_calc['A9'] = "D_Score"
    ws_calc['C9'] = '=C5*0.5 + (1-(Inputs!C3/1000))*0.3 + 0.2'
    
    # Hoja 4: Outputs
    ws_output = wb.create_sheet("Outputs")
    ws_output['A1'] = "OUTPUTS - HEALTH SCORE"
    ws_output['A1'].font = Font(bold=True, size=14)
    
    ws_output['A2'] = "Métrica"
    ws_output['C2'] = "Valor"
    ws_output['A2'].fill = HEADER_FILL
    ws_output['C2'].fill = HEADER_FILL
    
    ws_output['A3'] = "H_org"
    ws_output['C3'] = '=Parametros!C3*Calculos!C7 + Parametros!C4*Calculos!C8 + Parametros!C5*Calculos!C9'
    ws_output['C3'].number_format = '0.00'
    
    ws_output['A4'] = "Health_Band"
    ws_output['C4'] = '=IF(C3<0.35,"G1-Crítico",IF(C3<0.55,"G2-Bajo",IF(C3<0.75,"G3-Aceptable","G4-Saludable")))'
    
    ws_output['A6'] = "A_Score"
    ws_output['C6'] = '=Calculos!C7'
    ws_output['C6'].number_format = '0.00'
    
    ws_output['A7'] = "P_Score"
    ws_output['C7'] = '=Calculos!C8'
    ws_output['C7'].number_format = '0.00'
    
    ws_output['A8'] = "D_Score"
    ws_output['C8'] = '=Calculos!C9'
    ws_output['C8'].number_format = '0.00'
    
    # Hoja 5: Metadata
    ws_meta = wb.create_sheet("Metadata")
    ws_meta['A1'] = "METADATA Y TRAZABILIDAD"
    ws_meta['A1'].font = Font(bold=True, size=14)
    
    ws_meta['A3'] = "Version"
    ws_meta['C3'] = "1.0.0-CAP22"
    ws_meta['A4'] = "Fecha"
    ws_meta['C4'] = "2025-11-18"
    ws_meta['A5'] = "Fuente"
    ws_meta['C5'] = "CALCULADORAS_P0_SPEC.md § 3.1"
    ws_meta['A7'] = "Trazabilidad genoma:"
    ws_meta['A8'] = "- Teorema T6 (salud organizacional)"
    ws_meta['A9'] = "- Tejidos TF1/TF2/TF3"
    ws_meta['A10'] = "- Invariante I6 (convergencia)"
    ws_meta['A11'] = "- Primitivos P1/P2/P3"
    
    wb.save(output_path)
    print(f"✓ Generado: {output_path}")


def create_context_decision_matrix(output_path):
    """Genera context_decision_matrix.xlsx"""
    wb = Workbook()
    
    # Hoja 1: Context_Inputs
    ws_inputs = wb.active
    ws_inputs.title = "Context_Inputs"
    ws_inputs['A1'] = "CONTEXT DECISION MATRIX - INPUTS"
    ws_inputs['A1'].font = Font(bold=True, size=14)
    
    inputs_data = [
        ("Variable", "Valor", "Descripción"),
        ("H_org", 0.55, "Health score actual"),
        ("budget_total", 50000000, "Presupuesto total (CLP/USD)"),
        ("org_size", 50, "Tamaño organización"),
        ("context_risk", 3, "Nivel riesgo (1-5)"),
        ("strategic_horizon", 12, "Horizonte estratégico (meses)"),
        ("current_maturity", 2, "Madurez actual (1-5)"),
    ]
    
    for row_idx, row_data in enumerate(inputs_data, start=2):
        ws_inputs[f'A{row_idx}'] = row_data[0]
        ws_inputs[f'C{row_idx}'] = row_data[1]
        ws_inputs[f'D{row_idx}'] = row_data[2]
        if row_idx == 2:
            ws_inputs[f'A{row_idx}'].fill = HEADER_FILL
            ws_inputs[f'C{row_idx}'].fill = HEADER_FILL
            ws_inputs[f'D{row_idx}'].fill = HEADER_FILL
    
    # Hoja 2: Parametros
    ws_params = wb.create_sheet("Parametros")
    ws_params['A1'] = "PARAMETROS - UMBRALES DE TRAYECTORIA"
    ws_params['A1'].font = Font(bold=True, size=14)
    
    ws_params['A3'] = "Umbral_Survival"
    ws_params['C3'] = 0.35
    ws_params['D3'] = "H_org < 0.35 → Survival"
    
    ws_params['A4'] = "Umbral_Minimal"
    ws_params['C4'] = 0.55
    ws_params['D4'] = "H_org < 0.55 → Minimal"
    
    ws_params['A6'] = "Umbral_Avanzada"
    ws_params['C6'] = 0.75
    ws_params['D6'] = "H_org >= 0.75 → Avanzada"
    
    # Hoja 3: Decision_Rules
    ws_rules = wb.create_sheet("Decision_Rules")
    ws_rules['A1'] = "REGLAS DE DECISIÓN (DM1-DM5)"
    ws_rules['A1'].font = Font(bold=True, size=14)
    
    ws_rules['A3'] = "DM1: Si H_org < 0.35 → Survival"
    ws_rules['A4'] = "DM2: Si 0.35 ≤ H_org < 0.55 → Minimal"
    ws_rules['A5'] = "DM3: Si 0.55 ≤ H_org < 0.75 → Incremental"
    ws_rules['A6'] = "DM4: Si H_org ≥ 0.75 → Avanzada"
    ws_rules['A7'] = "DM5: Ajustar por riesgo y horizonte temporal"
    
    # Hoja 4: Trajectory_Output
    ws_output = wb.create_sheet("Trajectory_Output")
    ws_output['A1'] = "TRAYECTORIA SELECCIONADA"
    ws_output['A1'].font = Font(bold=True, size=14)
    
    ws_output['A2'] = "Métrica"
    ws_output['C2'] = "Valor"
    ws_output['A2'].fill = HEADER_FILL
    ws_output['C2'].fill = HEADER_FILL
    
    ws_output['A3'] = "trajectory_selected"
    ws_output['C3'] = '=IF(Context_Inputs!C3<Parametros!C4,"Survival",IF(Context_Inputs!C3<Parametros!C6,"Minimal","Avanzada"))'
    
    ws_output['A4'] = "rationale"
    ws_output['C4'] = '=IF(C3="Survival","H_org bajo o riesgo alto",IF(C3="Minimal","Condiciones intermedias","Alta madurez y recursos"))'
    
    ws_output['A5'] = "health_band_used"
    ws_output['C5'] = '=IF(Context_Inputs!C3<0.35,"G1",IF(Context_Inputs!C3<0.55,"G2",IF(Context_Inputs!C3<0.75,"G3","G4")))'
    
    # Hoja 5: Playbook_Mapping
    ws_playbook = wb.create_sheet("Playbook_Mapping")
    ws_playbook['A1'] = "MAPEO TRAYECTORIA → PLAYBOOKS"
    ws_playbook['A1'].font = Font(bold=True, size=14)
    
    ws_playbook['A3'] = "Survival"
    ws_playbook['C3'] = "P01-P03: Core stabilization"
    ws_playbook['A4'] = "Minimal"
    ws_playbook['C4'] = "P01-P06: Foundation building"
    ws_playbook['A5'] = "Incremental"
    ws_playbook['C5'] = "P04-P09: Capability expansion"
    ws_playbook['A6'] = "Avanzada"
    ws_playbook['C6'] = "P07-P12: Advanced optimization"
    
    # Hoja 6: Metadata
    ws_meta = wb.create_sheet("Metadata")
    ws_meta['A1'] = "METADATA Y TRAZABILIDAD"
    ws_meta['A1'].font = Font(bold=True, size=14)
    
    ws_meta['A3'] = "Version"
    ws_meta['C3'] = "1.0.0-CAP22"
    ws_meta['A4'] = "Fecha"
    ws_meta['C4'] = "2025-11-18"
    ws_meta['A5'] = "Fuente"
    ws_meta['C5'] = "CALCULADORAS_P0_SPEC.md § 3.2"
    ws_meta['A7'] = "Trazabilidad genoma:"
    ws_meta['A8'] = "- Axioma A4 (Propósito guía decisiones)"
    ws_meta['A9'] = "- Primitivos P4/P6 (Límite, Bandas contexto)"
    ws_meta['A10'] = "- Teorema T7 (Contexto determina trayectoria)"
    ws_meta['A11'] = "- Invariante I6 (Convergencia)"
    ws_meta['A12'] = "- Fases F1/F3 (Context → Trajectory)"
    
    wb.save(output_path)
    print(f"✓ Generado: {output_path}")


def create_convergence_tracker(output_path):
    """Genera convergence_tracker.xlsx"""
    wb = Workbook()
    
    # Hoja 1: E6_Current
    ws_current = wb.active
    ws_current.title = "E6_Current"
    ws_current['A1'] = "CONVERGENCE TRACKER - ESTADO ACTUAL (E6)"
    ws_current['A1'].font = Font(bold=True, size=14)
    
    current_data = [
        ("Variable", "Valor", "Descripción"),
        ("TF1_current_capacity_count", 5, "Capacidades desplegadas actuales"),
        ("TF2_current_flow_count", 3, "Flujos instrumentados actuales"),
        ("TF3_current_info_artifacts", 8, "Artefactos info actuales"),
        ("H_org_current", 0.55, "Health score actual"),
    ]
    
    for row_idx, row_data in enumerate(current_data, start=2):
        ws_current[f'A{row_idx}'] = row_data[0]
        ws_current[f'C{row_idx}'] = row_data[1]
        ws_current[f'D{row_idx}'] = row_data[2]
        if row_idx == 2:
            ws_current[f'A{row_idx}'].fill = HEADER_FILL
            ws_current[f'C{row_idx}'].fill = HEADER_FILL
            ws_current[f'D{row_idx}'].fill = HEADER_FILL
    
    # Hoja 2: E6_Target
    ws_target = wb.create_sheet("E6_Target")
    ws_target['A1'] = "ESTADO OBJETIVO (E6 Target)"
    ws_target['A1'].font = Font(bold=True, size=14)
    
    target_data = [
        ("Variable", "Valor", "Descripción"),
        ("TF1_target_capacity_count", 12, "Capacidades objetivo"),
        ("TF2_target_flow_count", 8, "Flujos objetivo"),
        ("TF3_target_info_artifacts", 20, "Artefactos info objetivo"),
        ("H_org_target", 0.75, "Health score objetivo"),
    ]
    
    for row_idx, row_data in enumerate(target_data, start=2):
        ws_target[f'A{row_idx}'] = row_data[0]
        ws_target[f'C{row_idx}'] = row_data[1]
        ws_target[f'D{row_idx}'] = row_data[2]
        if row_idx == 2:
            ws_target[f'A{row_idx}'].fill = HEADER_FILL
            ws_target[f'C{row_idx}'].fill = HEADER_FILL
            ws_target[f'D{row_idx}'].fill = HEADER_FILL
    
    # Hoja 3: Parametros
    ws_params = wb.create_sheet("Parametros")
    ws_params['A1'] = "PARAMETROS - PESOS CONVERGENCIA"
    ws_params['A1'].font = Font(bold=True, size=14)
    
    ws_params['A3'] = "w_TF1"
    ws_params['C3'] = 0.35
    ws_params['D3'] = "Peso tejido TF1"
    
    ws_params['A4'] = "w_TF2"
    ws_params['C4'] = 0.30
    ws_params['D4'] = "Peso tejido TF2"
    
    ws_params['A5'] = "w_TF3"
    ws_params['C5'] = 0.20
    ws_params['D5'] = "Peso tejido TF3"
    
    ws_params['A6'] = "w_H_org"
    ws_params['C6'] = 0.15
    ws_params['D6'] = "Peso H_org global"
    
    ws_params['A8'] = "convergence_threshold"
    ws_params['C8'] = 0.80
    ws_params['D8'] = "Umbral 'Converged'"
    
    # Hoja 4: Gap_Analysis
    ws_gap = wb.create_sheet("Gap_Analysis")
    ws_gap['A1'] = "ANÁLISIS DE GAP Y CONVERGENCIA"
    ws_gap['A1'].font = Font(bold=True, size=14)
    
    ws_gap['A2'] = "Métrica"
    ws_gap['C2'] = "Valor"
    ws_gap['A2'].fill = HEADER_FILL
    ws_gap['C2'].fill = HEADER_FILL
    
    ws_gap['A3'] = "TF1_convergence_ratio"
    ws_gap['C3'] = '=MIN(1, E6_Current!C3/E6_Target!C3)'
    ws_gap['C3'].number_format = '0.00'
    
    ws_gap['A4'] = "TF2_convergence_ratio"
    ws_gap['C4'] = '=MIN(1, E6_Current!C4/E6_Target!C4)'
    ws_gap['C4'].number_format = '0.00'
    
    ws_gap['A5'] = "TF3_convergence_ratio"
    ws_gap['C5'] = '=MIN(1, E6_Current!C5/E6_Target!C5)'
    ws_gap['C5'].number_format = '0.00'
    
    ws_gap['A6'] = "H_org_convergence_ratio"
    ws_gap['C6'] = '=MIN(1, E6_Current!C6/E6_Target!C6)'
    ws_gap['C6'].number_format = '0.00'
    
    ws_gap['A8'] = "Convergence_Score"
    ws_gap['C8'] = '=Parametros!C3*C3 + Parametros!C4*C4 + Parametros!C5*C5 + Parametros!C6*C6'
    ws_gap['C8'].number_format = '0.00'
    
    ws_gap['A9'] = "Convergence_Status"
    ws_gap['C9'] = '=IF(C8>=Parametros!C8,"Converged",IF(C8>=0.5,"In Progress","Lagging"))'
    
    # Hoja 5: Projections
    ws_proj = wb.create_sheet("Projections")
    ws_proj['A1'] = "PROYECCIONES TEMPORALES"
    ws_proj['A1'].font = Font(bold=True, size=14)
    
    ws_proj['A3'] = "Status"
    ws_proj['C3'] = '=Gap_Analysis!C9'
    
    ws_proj['A4'] = "Score actual"
    ws_proj['C4'] = '=Gap_Analysis!C8'
    ws_proj['C4'].number_format = '0.00'
    
    ws_proj['A5'] = "Gap restante"
    ws_proj['C5'] = '=1 - C4'
    ws_proj['C5'].number_format = '0.00'
    
    ws_proj['A7'] = "Estimación:"
    ws_proj['A8'] = "Tiempo estimado a convergencia (meses)"
    ws_proj['C8'] = '=IF(Gap_Analysis!C9="Converged",0,IF(Gap_Analysis!C9="In Progress",12,24))'
    
    # Hoja 6: Metadata
    ws_meta = wb.create_sheet("Metadata")
    ws_meta['A1'] = "METADATA Y TRAZABILIDAD"
    ws_meta['A1'].font = Font(bold=True, size=14)
    
    ws_meta['A3'] = "Version"
    ws_meta['C3'] = "1.0.0-CAP22"
    ws_meta['A4'] = "Fecha"
    ws_meta['C4'] = "2025-11-18"
    ws_meta['A5'] = "Fuente"
    ws_meta['C5'] = "CALCULADORAS_P0_SPEC.md § 3.3"
    ws_meta['A7'] = "Trazabilidad genoma:"
    ws_meta['A8'] = "- Primitivo P5 (Estado E6)"
    ws_meta['A9'] = "- Invariante I3 (Coherencia)"
    ws_meta['A10'] = "- Tejidos TF1/TF2/TF3"
    ws_meta['A11'] = "- Fases F9/F18 (Monitoring/Evolution)"
    
    wb.save(output_path)
    print(f"✓ Generado: {output_path}")


def main():
    """Genera las 3 calculadoras ORKO"""
    print("Generando calculadoras ORKO CAP-22...")
    print()
    
    # Rutas de salida
    base_dir = Path(__file__).parent.parent.parent / "calculadoras"
    base_dir.mkdir(exist_ok=True)
    
    health_path = base_dir / "health_score_calculator.xlsx"
    decision_path = base_dir / "context_decision_matrix.xlsx"
    convergence_path = base_dir / "convergence_tracker.xlsx"
    
    # Generar archivos
    create_health_score_calculator(health_path)
    create_context_decision_matrix(decision_path)
    create_convergence_tracker(convergence_path)
    
    print()
    print("=" * 60)
    print("✓ Generación completada exitosamente")
    print("=" * 60)
    print()
    print("Archivos generados:")
    print(f"  1. {health_path}")
    print(f"  2. {decision_path}")
    print(f"  3. {convergence_path}")
    print()
    print("Próximo paso: Validar archivos en Excel/LibreOffice")
    print("Siguiente: sq2 valida coherencia con CAP22_IMPL_GUIDE.md")


if __name__ == "__main__":
    main()
